from decimal import Decimal
from io import BytesIO
from xml.sax.saxutils import escape

from django.db.models import Q
from django.http import HttpResponse
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .models import ComputadorUsuario, Equipamento, ManutencaoEquipamento


COR_CABECALHO = "111827"
COR_SUBCABECALHO = "F1F5F9"
COR_BORDA = "CBD5E1"


def texto(valor, padrao="-"):
    if valor is None:
        return padrao

    valor = str(valor).strip()

    if not valor:
        return padrao

    return valor


def sim_nao(valor):
    return "Sim" if valor else "Não"


def data_br(valor):
    if not valor:
        return "-"

    return valor.strftime("%d/%m/%Y")


def data_hora_br(valor):
    if not valor:
        return "-"

    return timezone.localtime(valor).strftime("%d/%m/%Y %H:%M")


def moeda_br(valor):
    if valor is None or valor == "":
        return "-"

    try:
        numero = Decimal(valor)
    except Exception:
        return "-"

    texto_formatado = f"{numero:,.2f}"
    texto_formatado = texto_formatado.replace(",", "X").replace(".", ",").replace("X", ".")

    return f"R$ {texto_formatado}"


def nome_equipamento(equipamento):
    if not equipamento:
        return "-"

    partes = [
        equipamento.get_tipo_display(),
        equipamento.patrimonio or equipamento.numero_serie or "",
        equipamento.marca,
        equipamento.modelo,
    ]

    return " - ".join([texto_parte for texto_parte in partes if texto_parte])


def aplicar_estilo_excel(ws, titulo, total_colunas):
    ws.freeze_panes = "A4"

    ws.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=total_colunas,
    )

    celula_titulo = ws.cell(row=1, column=1)
    celula_titulo.value = titulo
    celula_titulo.font = Font(size=15, bold=True, color="FFFFFF")
    celula_titulo.fill = PatternFill("solid", fgColor=COR_CABECALHO)
    celula_titulo.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells(
        start_row=2,
        start_column=1,
        end_row=2,
        end_column=total_colunas,
    )

    celula_data = ws.cell(row=2, column=1)
    celula_data.value = f"Gerado em {timezone.localtime().strftime('%d/%m/%Y %H:%M')}"
    celula_data.font = Font(size=10, italic=True, color="475569")
    celula_data.fill = PatternFill("solid", fgColor="E2E8F0")
    celula_data.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    borda = Border(
        left=Side(style="thin", color=COR_BORDA),
        right=Side(style="thin", color=COR_BORDA),
        top=Side(style="thin", color=COR_BORDA),
        bottom=Side(style="thin", color=COR_BORDA),
    )

    for celula in ws[4]:
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = PatternFill("solid", fgColor=COR_CABECALHO)
        celula.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        celula.border = borda

    for linha in ws.iter_rows(min_row=5):
        for celula in linha:
            celula.alignment = Alignment(vertical="top", wrap_text=True)
            celula.border = borda

    for coluna in range(1, total_colunas + 1):
        letra = get_column_letter(coluna)
        largura = 14

        for celula in ws[letra]:
            valor = texto(celula.value, "")

            if len(valor) > largura:
                largura = len(valor)

        ws.column_dimensions[letra].width = min(max(largura + 2, 12), 38)


def resposta_excel(nome_arquivo, titulo, cabecalhos, linhas):
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatorio"

    ws.append([titulo])
    ws.append([f"Gerado em {timezone.localtime().strftime('%d/%m/%Y %H:%M')}"])
    ws.append([])
    ws.append(cabecalhos)

    for linha in linhas:
        ws.append(linha)

    aplicar_estilo_excel(ws, titulo, len(cabecalhos))

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{nome_arquivo}"'

    return response


def paragrafo_pdf(valor, estilo):
    return Paragraph(escape(texto(valor)), estilo)


def rodape_pdf(canvas, doc):
    canvas.saveState()
    canvas.setFont("Helvetica", 8)
    canvas.setFillColor(colors.HexColor("#64748B"))

    texto_rodape = f"Gerado em {timezone.localtime().strftime('%d/%m/%Y %H:%M')} - Pagina {canvas.getPageNumber()}"

    canvas.drawRightString(
        landscape(A4)[0] - 28,
        16,
        texto_rodape,
    )

    canvas.restoreState()


def resposta_pdf(nome_arquivo, titulo, subtitulo, cabecalhos, linhas, larguras):
    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=24,
        leftMargin=24,
        topMargin=24,
        bottomMargin=28,
    )

    styles = getSampleStyleSheet()

    estilo_titulo = ParagraphStyle(
        "TituloRelatorio",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=15,
        leading=18,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#111827"),
        spaceAfter=8,
    )

    estilo_subtitulo = ParagraphStyle(
        "SubtituloRelatorio",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9,
        leading=12,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#475569"),
        spaceAfter=14,
    )

    estilo_cabecalho = ParagraphStyle(
        "CabecalhoTabela",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=7,
        leading=9,
        alignment=TA_CENTER,
        textColor=colors.white,
    )

    estilo_celula = ParagraphStyle(
        "CelulaTabela",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=7,
        leading=9,
        alignment=TA_LEFT,
        textColor=colors.HexColor("#0F172A"),
    )

    elementos = [
        Paragraph(titulo, estilo_titulo),
        Paragraph(subtitulo, estilo_subtitulo),
        Spacer(1, 4),
    ]

    dados_tabela = [
        [paragrafo_pdf(cabecalho, estilo_cabecalho) for cabecalho in cabecalhos]
    ]

    for linha in linhas:
        dados_tabela.append([
            paragrafo_pdf(valor, estilo_celula)
            for valor in linha
        ])

    tabela = Table(
        dados_tabela,
        colWidths=larguras,
        repeatRows=1,
    )

    tabela.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111827")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))

    elementos.append(tabela)

    doc.build(elementos, onFirstPage=rodape_pdf, onLaterPages=rodape_pdf)

    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/pdf",
    )
    response["Content-Disposition"] = f'attachment; filename="{nome_arquivo}"'

    return response


def filtrar_equipamentos(request):
    busca = request.GET.get("q", "").strip()

    queryset = Equipamento.objects.select_related("setor").all()

    if busca:
        queryset = queryset.filter(
            Q(tipo__icontains=busca)
            | Q(patrimonio__icontains=busca)
            | Q(marca__icontains=busca)
            | Q(modelo__icontains=busca)
            | Q(numero_serie__icontains=busca)
            | Q(setor__nome__icontains=busca)
            | Q(usuario_responsavel__icontains=busca)
            | Q(status__icontains=busca)
            | Q(fornecedor__icontains=busca)
            | Q(numero_nota_fiscal__icontains=busca)
            | Q(origem__icontains=busca)
            | Q(observacoes__icontains=busca)
        ).distinct()

    return queryset.order_by("tipo", "marca", "modelo", "patrimonio")


def filtrar_computadores(request):
    busca = request.GET.get("q", "").strip()

    queryset = ComputadorUsuario.objects.select_related("setor").all()

    if busca:
        queryset = queryset.filter(
            Q(nome_usuario__icontains=busca)
            | Q(setor__nome__icontains=busca)
            | Q(ip_computador__icontains=busca)
            | Q(mac_address__icontains=busca)
            | Q(processador__icontains=busca)
            | Q(memoria_ram__icontains=busca)
            | Q(armazenamento_tipo__icontains=busca)
            | Q(armazenamento_capacidade__icontains=busca)
            | Q(observacoes__icontains=busca)
        )

    return queryset.order_by("nome_usuario")


def filtrar_manutencoes(request):
    busca = request.GET.get("q", "").strip()

    queryset = (
        ManutencaoEquipamento.objects
        .select_related("equipamento", "equipamento__setor")
        .all()
    )

    if busca:
        queryset = queryset.filter(
            Q(equipamento__patrimonio__icontains=busca)
            | Q(equipamento__marca__icontains=busca)
            | Q(equipamento__modelo__icontains=busca)
            | Q(equipamento__numero_serie__icontains=busca)
            | Q(equipamento__usuario_responsavel__icontains=busca)
            | Q(equipamento__setor__nome__icontains=busca)
            | Q(tipo_ocorrencia__icontains=busca)
            | Q(status__icontains=busca)
            | Q(responsavel_atendimento__icontains=busca)
            | Q(descricao__icontains=busca)
        )

    return queryset.order_by("-data_ocorrencia", "-criado_em")


def equipamentos_excel(request):
    cabecalhos = [
        "Tipo",
        "Patrimônio",
        "Marca",
        "Modelo",
        "Número de série",
        "Setor",
        "Responsável",
        "Status",
        "Produto novo",
        "Origem",
        "Data de compra",
        "Fornecedor",
        "Nota fiscal",
        "Valor de compra",
        "Garantia até",
        "Observações",
        "Criado em",
        "Atualizado em",
    ]

    linhas = []

    for equipamento in filtrar_equipamentos(request):
        linhas.append([
            equipamento.get_tipo_display(),
            equipamento.patrimonio or "-",
            equipamento.marca or "-",
            equipamento.modelo or "-",
            equipamento.numero_serie or "-",
            equipamento.setor.nome if equipamento.setor else "-",
            equipamento.usuario_responsavel or "-",
            equipamento.get_status_display(),
            sim_nao(equipamento.produto_novo),
            equipamento.get_origem_display() if equipamento.origem else "-",
            data_br(equipamento.data_compra),
            equipamento.fornecedor or "-",
            equipamento.numero_nota_fiscal or "-",
            moeda_br(equipamento.valor_compra),
            data_br(equipamento.garantia_ate),
            equipamento.observacoes or "-",
            data_hora_br(equipamento.criado_em),
            data_hora_br(equipamento.atualizado_em),
        ])

    return resposta_excel(
        "relatorio_equipamentos.xlsx",
        "Relatório de Equipamentos / Almoxarifado",
        cabecalhos,
        linhas,
    )


def equipamentos_pdf(request):
    cabecalhos = [
        "Tipo",
        "Patrimônio",
        "Marca / Modelo",
        "Setor",
        "Responsável",
        "Status",
        "Compra / Garantia",
    ]

    linhas = []

    for equipamento in filtrar_equipamentos(request):
        compra = []

        if equipamento.produto_novo:
            compra.append("Produto novo")

        if equipamento.origem:
            compra.append(equipamento.get_origem_display())

        if equipamento.valor_compra is not None:
            compra.append(moeda_br(equipamento.valor_compra))

        if equipamento.garantia_ate:
            compra.append(f"Garantia: {data_br(equipamento.garantia_ate)}")

        linhas.append([
            equipamento.get_tipo_display(),
            equipamento.patrimonio or "-",
            " ".join([parte for parte in [equipamento.marca, equipamento.modelo] if parte]) or "-",
            equipamento.setor.nome if equipamento.setor else "-",
            equipamento.usuario_responsavel or "-",
            equipamento.get_status_display(),
            " | ".join(compra) if compra else "-",
        ])

    return resposta_pdf(
        "relatorio_equipamentos.pdf",
        "Relatório de Equipamentos / Almoxarifado",
        "Resumo dos equipamentos cadastrados, status, setor, responsável e dados principais de compra.",
        cabecalhos,
        linhas,
        [62, 70, 130, 90, 100, 70, 170],
    )


def computadores_excel(request):
    cabecalhos = [
        "Usuário",
        "Setor",
        "IP",
        "MAC",
        "Usa especificações",
        "Processador",
        "Memória RAM",
        "Tipo armazenamento",
        "Capacidade armazenamento",
        "Fonte W",
        "Observações",
        "Criado em",
        "Atualizado em",
    ]

    linhas = []

    for computador in filtrar_computadores(request):
        linhas.append([
            computador.nome_usuario,
            computador.setor.nome if computador.setor else "-",
            computador.ip_computador,
            computador.mac_address,
            sim_nao(computador.mostrar_especificacoes),
            computador.processador or "-",
            computador.memoria_ram or "-",
            computador.get_armazenamento_tipo_display() if computador.armazenamento_tipo else "-",
            computador.armazenamento_capacidade or "-",
            f"{computador.fonte_watts}W" if computador.fonte_watts else "-",
            computador.observacoes or "-",
            data_hora_br(computador.criado_em),
            data_hora_br(computador.atualizado_em),
        ])

    return resposta_excel(
        "relatorio_computadores.xlsx",
        "Relatório de Computadores",
        cabecalhos,
        linhas,
    )


def computadores_pdf(request):
    cabecalhos = [
        "Usuário",
        "Setor",
        "IP",
        "MAC",
        "Especificações",
        "Observações",
    ]

    linhas = []

    for computador in filtrar_computadores(request):
        especificacoes = []

        if computador.processador:
            especificacoes.append(computador.processador)

        if computador.memoria_ram:
            especificacoes.append(computador.memoria_ram)

        if computador.armazenamento_tipo or computador.armazenamento_capacidade:
            especificacoes.append(
                " ".join([
                    parte for parte in [
                        computador.get_armazenamento_tipo_display() if computador.armazenamento_tipo else "",
                        computador.armazenamento_capacidade,
                    ]
                    if parte
                ])
            )

        if computador.fonte_watts:
            especificacoes.append(f"Fonte {computador.fonte_watts}W")

        linhas.append([
            computador.nome_usuario,
            computador.setor.nome if computador.setor else "-",
            computador.ip_computador,
            computador.mac_address,
            " | ".join(especificacoes) if especificacoes else "-",
            computador.observacoes or "-",
        ])

    return resposta_pdf(
        "relatorio_computadores.pdf",
        "Relatório de Computadores",
        "Resumo dos computadores cadastrados, rede e especificações principais.",
        cabecalhos,
        linhas,
        [100, 90, 80, 100, 190, 170],
    )


def manutencoes_excel(request):
    cabecalhos = [
        "Equipamento",
        "Setor",
        "Tipo de ocorrência",
        "Data da ocorrência",
        "Responsável",
        "Custo",
        "Status",
        "Descrição",
        "Criado em",
        "Atualizado em",
    ]

    linhas = []

    for manutencao in filtrar_manutencoes(request):
        linhas.append([
            nome_equipamento(manutencao.equipamento),
            manutencao.equipamento.setor.nome if manutencao.equipamento.setor else "-",
            manutencao.get_tipo_ocorrencia_display(),
            data_hora_br(manutencao.data_ocorrencia),
            manutencao.responsavel_atendimento or "-",
            moeda_br(manutencao.custo),
            manutencao.get_status_display(),
            manutencao.descricao or "-",
            data_hora_br(manutencao.criado_em),
            data_hora_br(manutencao.atualizado_em),
        ])

    return resposta_excel(
        "relatorio_historico_manutencoes.xlsx",
        "Relatório de Histórico / Manutenções",
        cabecalhos,
        linhas,
    )


def manutencoes_pdf(request):
    cabecalhos = [
        "Data",
        "Equipamento",
        "Ocorrência",
        "Responsável",
        "Status",
        "Custo",
        "Descrição",
    ]

    linhas = []

    for manutencao in filtrar_manutencoes(request):
        linhas.append([
            data_hora_br(manutencao.data_ocorrencia),
            nome_equipamento(manutencao.equipamento),
            manutencao.get_tipo_ocorrencia_display(),
            manutencao.responsavel_atendimento or "-",
            manutencao.get_status_display(),
            moeda_br(manutencao.custo),
            manutencao.descricao or "-",
        ])

    return resposta_pdf(
        "relatorio_historico_manutencoes.pdf",
        "Relatório de Histórico / Manutenções",
        "Resumo dos registros de manutenção, movimentação, limpeza, baixa e observações.",
        cabecalhos,
        linhas,
        [76, 145, 82, 85, 70, 62, 210],
    )